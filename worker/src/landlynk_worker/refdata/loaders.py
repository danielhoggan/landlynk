"""Server-side reference data loaders.

The worker downloads ONS/OS open data over HTTP and loads it straight into
PostGIS, so reference data is loaded from the deployed app with no local
commands. Boundaries come from the ONS Open Geography Portal ArcGIS service
(paginated GeoJSON); census and income come from CSV or XLSX URLs.

Each load replaces its target table and records provenance in reference_load in
one transaction, so a refresh is atomic and auditable.
"""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from collections.abc import Iterator
from typing import TYPE_CHECKING
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import httpx

from . import transforms as t

if TYPE_CHECKING:
    from psycopg_pool import ConnectionPool

_TIMEOUT = httpx.Timeout(120.0)


# --- HTTP download -----------------------------------------------------------


def _require_url(url: str) -> str:
    """Reject a blank URL up front with a clear message.

    Several datasets have no stable default and must be pasted; without this the
    underlying client raises an opaque "missing protocol" error.
    """
    if not url or not url.strip():
        raise ValueError("No URL provided. Paste the source URL and press Load.")
    return url.strip()


def _get_text(url: str) -> str:
    url = _require_url(url)
    with httpx.Client(timeout=_TIMEOUT, follow_redirects=True) as client:
        resp = client.get(url)
        resp.raise_for_status()
        return resp.text


def _get_bytes(url: str) -> bytes:
    url = _require_url(url)
    with httpx.Client(timeout=_TIMEOUT, follow_redirects=True) as client:
        resp = client.get(url)
        resp.raise_for_status()
        return resp.content


# ONS dataset pages link to the actual file as /file?uri=...xlsx (or xls/csv).
_ONS_FILE_RE = re.compile(
    r'href="(/file\?uri=[^"]+\.(?:xlsx|xlsm|xls|csv))"', re.IGNORECASE
)


def _resolve_data_url(url: str, data: bytes) -> str | None:
    """If `data` is an HTML page (e.g. an ONS dataset page was pasted instead of
    the file), find the first xlsx/xls/csv download link and return its absolute
    URL. Returns None when the content is already a data file."""
    head = data[:1024].lstrip().lower()
    if not (
        head.startswith(b"<!doctype html")
        or head.startswith(b"<html")
        or b"<head" in head
    ):
        return None
    from urllib.parse import urljoin

    match = _ONS_FILE_RE.search(data.decode("utf-8", "ignore"))
    if not match:
        raise ValueError(
            "That URL is a web page, not a data file. Open the dataset page, then "
            "copy the link of the xlsx or csv download itself and paste that."
        )
    return urljoin(url, match.group(1))


def fetch_arcgis_geojson(
    query_url: str,
    out_fields: str = "*",
    page_size: int = 2000,
    client: httpx.Client | None = None,
) -> dict:
    """Page an ArcGIS FeatureServer query endpoint into one GeoJSON FeatureCollection.

    ``query_url`` is the layer's .../query URL (from the ONS portal API). Our
    parameters are merged over any already present so a pasted URL still works.
    The client is injectable for testing.
    """
    parts = urlparse(_require_url(query_url))
    base_params = dict(parse_qsl(parts.query))
    owned = client is None
    client = client or httpx.Client(timeout=_TIMEOUT, follow_redirects=True)
    features: list[dict] = []
    offset = 0
    # Advance by the number actually returned, not the requested page size:
    # ArcGIS servers commonly cap a page below the requested count, which would
    # otherwise skip records. Stop when a page comes back empty.
    try:
        for _ in range(10_000):  # safety bound against a server ignoring offset
            params = {
                **base_params,
                "where": base_params.get("where", "1=1"),
                "outFields": out_fields,
                "outSR": "4326",
                "f": "geojson",
                "resultOffset": str(offset),
                "resultRecordCount": str(page_size),
            }
            url = urlunparse(parts._replace(query=urlencode(params)))
            resp = client.get(url)
            resp.raise_for_status()
            batch = resp.json().get("features") or []
            if not batch:
                break
            features.extend(batch)
            offset += len(batch)
    finally:
        if owned:
            client.close()
    return {"type": "FeatureCollection", "features": features}


def _csv_rows(text: str) -> list[dict]:
    """Parse CSV text into dict rows, handling a BOM and the delimiter. Does not
    require an area-code column, so point files (schools, hospitals) parse too."""
    if text.startswith("﻿"):
        text = text[1:]
    try:
        delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t").delimiter
    except csv.Error:
        delimiter = ","
    return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))


def _read_csv(text: str) -> tuple[list[dict], str]:
    rows = _csv_rows(text)
    fieldnames = list(rows[0].keys()) if rows else []
    code_field = t.find_area_code_field(fieldnames)
    return rows, code_field


# Keyword to find the right geography CSV inside a NOMIS bulk zip.
_AREA_TYPE_KEYWORD = {"MSOA": "msoa", "LA": "ltla"}


def _fetch_csv_text(url: str, area_type: str) -> str:
    """Return CSV text from a URL, transparently handling a NOMIS bulk .zip.

    NOMIS bulk census downloads are zips containing one CSV per geography
    (e.g. census2021-ts054-msoa.csv). We pick the member matching the area type
    so census loads are a one-click download.
    """
    if not url.lower().endswith(".zip"):
        return _get_text(url)
    content = _get_bytes(url)
    keyword = _AREA_TYPE_KEYWORD.get(area_type, "msoa")
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
        match = next((n for n in csv_names if keyword in n.lower()), None)
        if match is None and area_type == "LA":
            # Fall back to other LA spellings used across releases.
            match = next(
                (n for n in csv_names if any(k in n.lower() for k in ("lad", "utla"))),
                None,
            )
        if match is None and csv_names:
            match = csv_names[0]
        if match is None:
            raise ValueError(f"No CSV found in zip at {url}")
        return zf.read(match).decode("utf-8-sig")


_XLSX_CODE_HINTS = (
    "msoa code",
    "area code",
    "geography code",
    "msoa11cd",
    "msoa21cd",
    "lad code",
    "la code",
)


def _looks_like_header(cells: list[str]) -> bool:
    """True if a row looks like the data header (carries an area-code column).

    ONS spreadsheets have title and notes rows before the real header, so we
    scan for the first row that has an area-code column and several fields.
    """
    low = [c.lower() for c in cells]
    has_code = any(c in _XLSX_CODE_HINTS for c in low) or any(
        "msoa" in c and "code" in c for c in low
    )
    return has_code and sum(1 for c in cells if c) >= 3


def _read_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:40]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if _looks_like_header(cells):
                records = [dict(zip(cells, r, strict=False)) for r in rows[i + 1 :]]
                records = [r for r in records if any(v is not None for v in r.values())]
                if records:
                    return records
    # Fallback: first sheet, header in row 1.
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(it)]
    return [dict(zip(header, v, strict=False)) for v in it]


def _read_xlsx_sheets(content: bytes) -> list[list[dict]]:
    """Every worksheet as a record list, for workbooks whose right sheet must be
    discovered by its columns (e.g. the ONS green space file has separate sheets
    per geography and per measure). Only sheets with a recognisable data header
    are returned, so callers can try each until the columns they need appear.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out: list[list[dict]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:40]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if _looks_like_header(cells):
                header, data_start = _resolve_header(cells, rows, i)
                records = [
                    dict(zip(header, r, strict=False)) for r in rows[data_start:]
                ]
                records = [r for r in records if any(v is not None for v in r.values())]
                if records:
                    out.append(records)
                break
    return out


def _resolve_header(
    cells: list[str], rows: list, i: int
) -> tuple[list[str], int]:
    """Resolve a possibly two-row header into one label list and the data start.

    ONS reference tables put geography codes on the header row and the measure
    names on the row beneath (the geography cells there are blank). When that
    layout is detected, the two rows are merged so the measure columns get their
    names; otherwise the single header row is used.
    """
    code_idx = next(
        (j for j, c in enumerate(cells) if "code" in c.lower()), 0
    )
    if i + 1 < len(rows):
        nxt = [str(c).strip() if c is not None else "" for c in rows[i + 1]]
        # A sub-header row leaves the area-code column blank but names later
        # (measure) columns; a real data row fills the area-code column.
        code_blank = code_idx >= len(nxt) or not nxt[code_idx]
        names_later = any(nxt[j] for j in range(code_idx + 1, len(nxt)))
        if code_blank and names_later:
            merged = [
                (nxt[j] if j < len(nxt) and nxt[j] else (cells[j] if j < len(cells) else ""))
                for j in range(max(len(cells), len(nxt)))
            ]
            return merged, i + 2
    return cells, i + 1


def _load_tabular_sheets(url: str) -> list[list[dict]]:
    """Load a file as one record list per sheet (CSV/XLS yield a single sheet),
    following an ONS dataset page link if a page URL was pasted."""
    data = _get_bytes(url)
    resolved = _resolve_data_url(url, data)
    if resolved:
        url = resolved
        data = _get_bytes(url)
    return _tabular_sheets_from_bytes(url, data)


def _tabular_sheets_from_bytes(name: str, data: bytes) -> list[list[dict]]:
    """One record list per sheet from raw bytes, by file extension. xlsx yields
    every sheet (for workbooks whose right sheet is found by its columns)."""
    low = name.lower()
    if low.endswith(".xls"):
        return [_read_xls(data)]
    if low.endswith((".xlsx", ".xlsm")):
        return _read_xlsx_sheets(data)
    return [_csv_rows(data.decode("utf-8-sig", "ignore"))]


def _read_xls(content: bytes) -> list[dict]:
    """Read a legacy .xls workbook (e.g. ONS HPSSA) into records.

    Mirrors _read_xlsx: find the header row by its shape, then map rows to dicts.
    """
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    for sheet in wb.sheets():
        rows = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        for i, row in enumerate(rows[:40]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if _looks_like_header(cells):
                records = [dict(zip(cells, r, strict=False)) for r in rows[i + 1 :]]
                records = [r for r in records if any(v != "" for v in r.values())]
                if records:
                    return records
    sheet = wb.sheet_by_index(0)
    header = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
    return [
        dict(
            zip(
                header,
                [sheet.cell_value(r, c) for c in range(sheet.ncols)],
                strict=False,
            )
        )
        for r in range(1, sheet.nrows)
    ]


# --- Transforms to rows ------------------------------------------------------

_BOUNDARY_KEYS = {
    "MSOA": {"code": "MSOA21CD", "name": "MSOA21NM"},
    "LA": {"code": "LAD21CD", "name": "LAD21NM"},
}


def _boundary_rows(geojson: dict, area_type: str) -> list[dict]:
    keys = _BOUNDARY_KEYS[area_type]
    rows: list[dict] = []
    for feature in geojson.get("features") or []:
        props = feature.get("properties") or {}
        geometry = feature.get("geometry")
        code = str(props.get(keys["code"]) or "").strip()
        if not code or geometry is None:
            continue
        rows.append(
            {
                "area_code": code,
                "area_type": area_type,
                "area_name": str(props.get(keys["name"]) or "").strip() or None,
                "geom": json.dumps(geometry),
            }
        )
    return rows


def _demographics_rows(
    age_text: str, households_text: str, area_type: str
) -> list[dict]:
    age_rows, age_code = _read_csv(age_text)
    hh_rows, hh_code = _read_csv(households_text)

    households_by_area: dict[str, tuple[int | None, float | None]] = {}
    for record in hh_rows:
        code = (record.get(hh_code) or "").strip()
        if not code:
            continue
        # TS003 total reads 'Household composition: Total'; family is the single
        # family household aggregate (not its sub-rows, which would double count).
        total = t.category_value(record, hh_code, "total", "all households")
        family = t.category_value(
            record, hh_code, "single family household", "one family household"
        )
        households_by_area[code] = (total, t.share(family, total))

    rows: list[dict] = []
    for record in age_rows:
        code = (record.get(age_code) or "").strip()
        if not code:
            continue
        counts = t.single_year_counts(record, age_code)
        bands = t.aggregate_age_bands(counts)
        population = sum(v for v in counts.values() if v is not None) or None
        households, family_share = households_by_area.get(code, (None, None))
        rows.append(
            {
                "area_code": code,
                "area_type": area_type,
                "population": population,
                "households": households,
                **bands,
                "median_age": t.median_age(counts),
                "family_household_share": family_share,
            }
        )
    return rows


def _tenure_rows(text: str, area_type: str) -> list[dict]:
    records, code_field = _read_csv(text)
    rows: list[dict] = []
    for record in records:
        code = (record.get(code_field) or "").strip()
        if not code:
            continue
        # TS054 total reads 'Tenure of household: Total'. Each tenure is the
        # category aggregate, matched exactly so the aggregate is not summed with
        # its own sub-rows (e.g. social rented and its council/other breakdown).
        total = t.category_value(record, code_field, "total", "all households")
        rows.append(
            {
                "area_code": code,
                "area_type": area_type,
                "owns_outright": t.share(
                    t.category_value(record, code_field, "owns outright"), total
                ),
                "owns_with_mortgage": t.share(
                    t.category_value(
                        record,
                        code_field,
                        "owns with a mortgage or loan",
                        "owns with a mortgage or shared ownership",
                        "shared ownership",
                    ),
                    total,
                ),
                "social_rented": t.share(
                    t.category_value(record, code_field, "social rented"), total
                ),
                "private_rented": t.share(
                    t.category_value(
                        record, code_field, "private rented", "lives rent free"
                    ),
                    total,
                ),
            }
        )
    return rows


def _income_rows(records: list[dict], area_type: str) -> list[dict]:
    if not records:
        return []
    fieldnames = list(records[0].keys())
    code_col = t.find_column(
        fieldnames, ("msoa code", "area code", "geography code", "code")
    )
    # ONS small-area income is mean-based ("Net annual income (£)"); match it
    # broadly. Median is often absent at MSOA level.
    mean_col = t.find_column(
        fieldnames,
        (
            "net annual income (mean)",
            "net annual income",
            "mean income",
            "mean",
            "income",
        ),
    )
    median_col = t.find_column(
        fieldnames, ("net annual income (median)", "median income", "median")
    )
    if code_col is None:
        raise ValueError(f"No area code column found in {fieldnames}")
    rows: list[dict] = []
    for record in records:
        code = str(record.get(code_col) or "").strip()
        if not code:
            continue
        rows.append(
            {
                "area_code": code,
                "area_type": area_type,
                "median_income": (
                    t.parse_number(record.get(median_col)) if median_col else None
                ),
                "mean_income": (
                    t.parse_number(record.get(mean_col)) if mean_col else None
                ),
            }
        )
    return rows


def _house_price_rows(records: list[dict], area_type: str) -> list[dict]:
    """ONS HPSSA median price paid by area, taking the most recent period.

    HPSSA sheets carry one column per rolling year, running chronologically from
    1995 (e.g. "Year ending Dec 1995" ... "Year ending Sep 2023"). We must take
    the most recent (rightmost dated) column, not the first match, or prices read
    decades out of date.
    """
    if not records:
        return []
    fieldnames = list(records[0].keys())
    code_col = t.find_column(
        fieldnames, ("msoa code", "area code", "geography code", "code")
    )
    if code_col is None:
        raise ValueError(f"No area code column found in {fieldnames}")
    # A single explicit median price column (some CSV variants), else the latest
    # dated period column. Dated columns are ordered oldest to newest left to
    # right, so the last one holding a 19xx/20xx year is the most recent price.
    price_col = t.find_column(fieldnames, ("median price", "median_price"))
    if price_col is None:
        dated = [f for f in fieldnames if re.search(r"(?:19|20)\d{2}", str(f))]
        price_col = dated[-1] if dated else t.find_column(fieldnames, ("price",))
    rows: list[dict] = []
    for record in records:
        code = str(record.get(code_col) or "").strip()
        if not code:
            continue
        rows.append(
            {
                "area_code": code,
                "area_type": area_type,
                "median_price": (
                    t.parse_number(record.get(price_col)) if price_col else None
                ),
            }
        )
    return rows


# --- DB write ----------------------------------------------------------------


def _records_from_bytes(url: str, data: bytes) -> list[dict]:
    """Parse already-downloaded bytes, following an ONS page link if pasted."""
    resolved = _resolve_data_url(url, data)
    if resolved:
        url = resolved
        data = _get_bytes(url)
    low = url.lower()
    if low.endswith(".xls"):
        return _read_xls(data)
    if low.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(data)
    return _csv_rows(data.decode("utf-8-sig", "ignore"))


def _load_records(url: str) -> list[dict]:
    """Fetch a tabular file (xlsx/xls/csv), following an ONS page link if pasted."""
    return _records_from_bytes(url, _get_bytes(url))


def _load_records_dated(url: str, max_back: int = 10) -> list[dict]:
    """Load a date-stamped file (e.g. GIAS edubasealldataYYYYMMDD.csv).

    Files like GIAS are published daily with the date in the name, so there is no
    fixed link. Given a URL containing a YYYYMMDD, try that date then step back a
    day at a time until one downloads, so a default of today's date just works
    even before the day's file is published. Falls back to a clear error asking
    for a manual URL if none in the window exist.
    """
    from datetime import datetime, timedelta

    match = re.search(r"\d{8}", url)
    if not match:
        return _load_records(url)
    stamp = match.group(0)
    try:
        base = datetime.strptime(stamp, "%Y%m%d").date()
    except ValueError:
        return _load_records(url)
    last_error: Exception | None = None
    for i in range(max_back):
        day = (base - timedelta(days=i)).strftime("%Y%m%d")
        candidate = url.replace(stamp, day, 1)
        try:
            return _records_from_bytes(candidate, _get_bytes(candidate))
        except Exception as exc:  # download miss for that day; try the day before
            last_error = exc
    raise ValueError(
        "Could not find a published file in the last "
        f"{max_back} days from {url}. Paste a specific file URL instead. "
        f"Last error: {last_error}"
    )


def _is_msoa(code: str) -> bool:
    return code[:3] in ("E02", "W02")


def _green_space_rows(records: list[dict], area_type: str) -> list[dict]:
    """ONS access to green space, as minutes walk to the nearest green space.

    The ONS public green space tables are published at LSOA, but each row carries
    its MSOA code, so we mean the distance over the LSOAs in each MSOA. Distance
    (metres) to the nearest park or public green space is converted to a walk
    time at ~80 m/min. A file that is already MSOA-level (one row per MSOA) falls
    out of the same mean. Non-MSOA rows are ignored.
    """
    if not records:
        return []
    fields = list(records[0].keys())
    code_col = t.find_column(fields, ("msoa code", "msoa11cd", "msoa21cd"))
    if code_col is None:
        code_col = t.find_column(fields, ("area code", "geography code"))
    dist_col = t.find_column(
        fields,
        (
            "average distance to nearest park",
            "average distance to nearest publicly",
            "distance to nearest park",
            "distance to nearest green",
            "distance to nearest public",
            "distance to nearest",
            "average distance",
        ),
    )
    # Return empty when this sheet has no distance column, so a caller scanning a
    # multi-sheet workbook can move on to the next sheet.
    if code_col is None or dist_col is None:
        return []
    by_msoa: dict[str, list[float]] = {}
    for r in records:
        code = str(r.get(code_col) or "").strip()
        if not _is_msoa(code):
            continue
        metres = t.parse_number(r.get(dist_col))
        if metres is None:
            continue
        by_msoa.setdefault(code, []).append(metres)
    return [
        {
            "area_code": code,
            "area_type": area_type,
            "metric_key": "greenspace_minutes",
            "value": round(_mean(metres_list) / 80.0, 1),
        }
        for code, metres_list in by_msoa.items()
    ]


def _imd_rows(records: list[dict], lookup: list[dict], area_type: str) -> list[dict]:
    """Index of Multiple Deprivation, aggregated from LSOA to MSOA.

    IMD is published at LSOA, so we map LSOA to MSOA with the supplied lookup and
    take the mean score and mean decile per MSOA (an indicative area summary).
    """
    if not records or not lookup:
        return []
    lf = list(lookup[0].keys())
    lsoa_l = t.find_column(lf, ("lsoa code", "lsoa11cd", "lsoa21cd", "lsoa"))
    msoa_l = t.find_column(lf, ("msoa code", "msoa11cd", "msoa21cd", "msoa"))
    if lsoa_l is None or msoa_l is None:
        raise ValueError(f"Lookup needs LSOA and MSOA code columns, got {lf}")
    to_msoa = {
        str(r.get(lsoa_l) or "").strip(): str(r.get(msoa_l) or "").strip()
        for r in lookup
    }

    rf = list(records[0].keys())
    lsoa_c = t.find_column(rf, ("lsoa code", "lsoa11cd", "lsoa21cd", "lsoa"))
    score_c = t.find_column(
        rf,
        (
            "index of multiple deprivation (imd) score",
            "index of multiple deprivation score",
            "imd score",
        ),
    )
    decile_c = t.find_column(rf, ("decile",))
    if lsoa_c is None or (score_c is None and decile_c is None):
        raise ValueError(f"No LSOA code / IMD score column found in {rf}")

    scores: dict[str, list[float]] = {}
    deciles: dict[str, list[float]] = {}
    for r in records:
        msoa = to_msoa.get(str(r.get(lsoa_c) or "").strip())
        if not _is_msoa(msoa or ""):
            continue
        if score_c:
            s = t.parse_number(r.get(score_c))
            if s is not None:
                scores.setdefault(msoa, []).append(s)
        if decile_c:
            d = t.parse_number(r.get(decile_c))
            if d is not None:
                deciles.setdefault(msoa, []).append(d)

    rows: list[dict] = []
    for msoa in set(scores) | set(deciles):
        if scores.get(msoa):
            rows.append(_metric_row(msoa, area_type, "imd_score", _mean(scores[msoa])))
        if deciles.get(msoa):
            rows.append(
                _metric_row(msoa, area_type, "imd_decile", round(_mean(deciles[msoa])))
            )
    return rows


def _mean(values: list[float]) -> float:
    return sum(values) / len(values)


def _metric_row(code: str, area_type: str, key: str, value: float) -> dict:
    return {
        "area_code": code,
        "area_type": area_type,
        "metric_key": key,
        "value": value,
    }


def _replace_metric(
    pool: ConnectionPool, metric_keys: list[str], rows: list[dict]
) -> int:
    """Replace all rows for the given metric_keys with the supplied rows."""
    with pool.connection() as conn, conn.transaction():
        conn.execute(
            "DELETE FROM area_metric WHERE metric_key = ANY(%s)", [list(metric_keys)]
        )
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT INTO area_metric (area_code, area_type, metric_key, value) "
                "VALUES (%s, %s, %s, %s)",
                [
                    (r["area_code"], r["area_type"], r["metric_key"], r["value"])
                    for r in rows
                ],
            )
    return len(rows)


def load_green_space(pool: ConnectionPool, url: str, area_type: str = "MSOA") -> int:
    return _load_green_space_sheets(pool, _load_tabular_sheets(url), area_type)


def load_green_space_bytes(
    pool: ConnectionPool, name: str, data: bytes, area_type: str = "MSOA"
) -> int:
    """Load green space from an uploaded workbook (the ONS public green space
    reference tables)."""
    resolved = _resolve_data_url(name, data)
    if resolved:
        data = _get_bytes(resolved)
        name = resolved
    return _load_green_space_sheets(
        pool, _tabular_sheets_from_bytes(name, data), area_type
    )


def _load_green_space_sheets(
    pool: ConnectionPool, sheets: list[list[dict]], area_type: str
) -> int:
    # The ONS green space workbook has many sheets (per geography and measure);
    # scan them for the one carrying an MSOA code and a distance column.
    rows: list[dict] = []
    for records in sheets:
        rows = _green_space_rows(records, area_type)
        if rows:
            break
    if not rows:
        raise ValueError(
            "No MSOA distance-to-green-space sheet found. This looks like the "
            "private outdoor space (gardens) file, which has no green space "
            "distance. Use the ONS public green space reference tables instead."
        )
    return _replace_metric(pool, ["greenspace_minutes"], rows)


def load_imd(
    pool: ConnectionPool,
    url: str,
    lookup_url: str | None = None,
    area_type: str = "MSOA",
) -> int:
    # The IMD file is LSOA-level; map to MSOA with a pasted lookup, or fall back
    # to the lookup built from the postcode directory when Postcodes is loaded.
    if lookup_url and lookup_url.strip():
        lookup = _load_records(lookup_url)
    else:
        lookup = _lsoa_msoa_lookup(pool)
        if not lookup:
            raise ValueError(
                "No LSOA to MSOA lookup available. Load the Postcodes dataset "
                "first (it builds the lookup), or paste a lookup CSV URL."
            )
    rows = _imd_rows(_load_records(url), lookup, area_type)
    return _replace_metric(pool, ["imd_score", "imd_decile"], rows)


def _lsoa_msoa_lookup(
    pool: ConnectionPool,
) -> list[dict]:  # pragma: no cover - needs DB
    """LSOA to MSOA pairs from the table built during the postcode load.

    Returns an empty list if the table is missing (migration not yet applied) or
    empty, so the caller can give a friendly "load Postcodes first" message
    rather than surfacing a raw SQL error.
    """
    try:
        with pool.connection() as conn:
            rows = conn.execute("SELECT lsoa, msoa FROM lsoa_msoa").fetchall()
    except Exception:
        return []
    return [{"lsoa": lsoa, "msoa": msoa} for lsoa, msoa in rows]


# --- point datasets (schools, crime): point-in-MSOA via PostGIS ---------------


def _bng_to_wgs(easting: float, northing: float) -> tuple[float, float]:
    from pyproj import Transformer

    global _BNG
    try:
        tr = _BNG
    except NameError:
        tr = _BNG = Transformer.from_crs("EPSG:27700", "EPSG:4326", always_xy=True)
    lng, lat = tr.transform(easting, northing)
    return lng, lat


_GOOD_RATINGS = {"outstanding", "good"}
_POOR_RATINGS = {
    "requires improvement",
    "inadequate",
    "serious weaknesses",
    "special measures",
}


def _ofsted_good_by_urn(records: list[dict]) -> dict[str, bool]:
    """Map school URN to a Good-or-Outstanding flag from an Ofsted outcomes file.

    Ofsted ratings are published separately from GIAS, keyed by URN. The outcome
    is sometimes text (Good) and sometimes a code (1 Outstanding, 2 Good).
    """
    if not records:
        return {}
    f = list(records[0].keys())
    urn_c = t.find_column(f, ("urn",))
    rate_c = t.find_column(
        f,
        (
            "overall effectiveness",
            "ofsted rating",
            "ofstedrating",
            "outcome",
        ),
    )
    if urn_c is None or rate_c is None:
        raise ValueError(f"Ofsted file needs URN and an outcome/rating column, got {f}")
    out: dict[str, bool] = {}
    for r in records:
        urn = str(r.get(urn_c) or "").strip()
        if not urn:
            continue
        raw = str(r.get(rate_c) or "").strip().lower()
        good: bool | None = None
        if raw in _GOOD_RATINGS:
            good = True
        elif raw in _POOR_RATINGS:
            good = False
        else:
            num = t.parse_number(raw)
            if num is not None:
                good = int(num) in (1, 2)  # 1 Outstanding, 2 Good
        if good is not None:
            out[urn] = good
    return out


def _school_points(
    records: list[dict], good_by_urn: dict[str, bool] | None = None
) -> list[dict]:
    """Open schools as {lng, lat, good} from a GIAS extract.

    The standard GIAS extract carries the location (easting/northing, or
    lat/long) but not the Ofsted rating, so ``good`` is taken from a separate
    Ofsted file by URN when supplied and is None otherwise. Only open schools
    with a location are kept; rating is optional.
    """
    if not records:
        return []
    f = list(records[0].keys())
    east_c = t.find_column(f, ("easting",))
    north_c = t.find_column(f, ("northing",))
    lat_c = t.find_column(f, ("latitude", "lat"))
    lng_c = t.find_column(f, ("longitude", "long", "lng"))
    urn_c = t.find_column(f, ("urn",))
    rating_c = t.find_column(
        f, ("ofstedrating (name)", "ofsted rating", "ofstedrating")
    )
    status_c = t.find_column(f, ("establishmentstatus (name)", "establishmentstatus"))
    if (east_c is None or north_c is None) and (lat_c is None or lng_c is None):
        raise ValueError(
            f"Schools file needs easting/northing or latitude/longitude, got {f}"
        )
    points: list[dict] = []
    for r in records:
        if status_c and not str(r.get(status_c) or "").strip().lower().startswith(
            "open"
        ):
            continue
        lat = lng = None
        if lat_c and lng_c:
            lat = t.parse_number(r.get(lat_c))
            lng = t.parse_number(r.get(lng_c))
        if (lat is None or lng is None) and east_c and north_c:
            easting = t.parse_number(r.get(east_c))
            northing = t.parse_number(r.get(north_c))
            if easting and northing:
                lng, lat = _bng_to_wgs(easting, northing)
        if lat is None or lng is None or (lat == 0 and lng == 0):
            continue
        good: bool | None = None
        urn = str(r.get(urn_c) or "").strip() if urn_c else ""
        if good_by_urn and urn in good_by_urn:
            good = good_by_urn[urn]
        elif rating_c:
            rating = str(r.get(rating_c) or "").strip().lower()
            if rating and rating not in ("not applicable", "", "none"):
                good = rating in _GOOD_RATINGS
        points.append({"lng": lng, "lat": lat, "good": good})
    return points


def _crime_points(records: list[dict]) -> list[tuple[float, float]]:
    """Crime incidents as (lng, lat) from data.police.uk street-level CSV rows."""
    if not records:
        return []
    f = list(records[0].keys())
    lng_c = t.find_column(f, ("longitude", "long", "lng"))
    lat_c = t.find_column(f, ("latitude", "lat"))
    if lng_c is None or lat_c is None:
        raise ValueError(f"Crime CSV needs longitude and latitude, got {f}")
    out: list[tuple[float, float]] = []
    for r in records:
        lng = t.parse_number(r.get(lng_c))
        lat = t.parse_number(r.get(lat_c))
        if lng is not None and lat is not None:
            out.append((lng, lat))
    return out


def load_schools(
    pool: ConnectionPool,
    url: str,
    ratings_url: str | None = None,
    area_type: str = "MSOA",
) -> int:  # pragma: no cover - PostGIS spatial join, exercised live
    good_by_urn = (
        _ofsted_good_by_urn(_load_records(ratings_url)) if ratings_url else {}
    )
    points = _school_points(_load_records_dated(url), good_by_urn)
    if not points:
        return 0
    with pool.connection() as conn, conn.transaction():
        conn.execute(
            "CREATE TEMP TABLE _sch (lng float8, lat float8, good bool) "
            "ON COMMIT DROP"
        )
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT INTO _sch VALUES (%s, %s, %s)",
                [(p["lng"], p["lat"], p["good"]) for p in points],
            )
        conn.execute(
            "DELETE FROM area_metric WHERE metric_key IN "
            "('schools_count', 'schools_good_pct')"
        )
        conn.execute(
            "INSERT INTO area_metric (area_code, area_type, metric_key, value) "
            "SELECT b.area_code, %s, 'schools_count', count(*) "
            "FROM _sch s JOIN geo_boundaries b "
            "ON b.area_type = %s AND ST_Contains("
            "b.geom, ST_SetSRID(ST_MakePoint(s.lng, s.lat), 4326)) "
            "GROUP BY b.area_code",
            [area_type, area_type],
        )
        # Good-or-Outstanding share, only where ratings are known: count rated
        # schools as the denominator and skip areas with none rated, so the
        # metric is absent rather than misleadingly zero when no Ofsted file
        # was supplied.
        conn.execute(
            "INSERT INTO area_metric (area_code, area_type, metric_key, value) "
            "SELECT b.area_code, %s, 'schools_good_pct', "
            "round(100.0 * count(*) FILTER (WHERE s.good) "
            "/ NULLIF(count(*) FILTER (WHERE s.good IS NOT NULL), 0)) "
            "FROM _sch s JOIN geo_boundaries b "
            "ON b.area_type = %s AND ST_Contains("
            "b.geom, ST_SetSRID(ST_MakePoint(s.lng, s.lat), 4326)) "
            "GROUP BY b.area_code "
            "HAVING count(*) FILTER (WHERE s.good IS NOT NULL) > 0",
            [area_type, area_type],
        )
        n = conn.execute(
            "SELECT count(*) FROM area_metric WHERE metric_key = 'schools_count'"
        ).fetchone()[0]
    return int(n)


# --- postcodes (ONS Postcode Directory) and ODS hospital geocoding ------------


def _norm_postcode(value: str) -> str:
    """Normalise a postcode for keyed lookup: uppercase, no spaces."""
    return re.sub(r"\s+", "", value or "").upper()


def _postcode_centroid(row: dict) -> tuple[str, float, float] | None:
    """One ONSPD/NSPL row to (normalised postcode, lat, lng), or None.

    Terminated postcodes with no grid reference are published with a sentinel
    latitude of 99.999999; those and any (0, 0) are dropped.
    """
    pc_col = (
        row.get("pcds") or row.get("pcd") or row.get("Postcode") or row.get("postcode")
    )
    pc = _norm_postcode(str(pc_col or ""))
    if not pc:
        return None
    lat = t.parse_number(row.get("lat") or row.get("Latitude"))
    lng = t.parse_number(row.get("long") or row.get("Longitude"))
    if lat is None or lng is None or lat > 90 or (lat == 0 and lng == 0):
        return None
    return pc, lat, lng


def _onspd_csv_member(zf: zipfile.ZipFile) -> str:
    """The big postcode CSV inside an ONSPD/NSPL zip (under Data/)."""
    names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
    preferred = [n for n in names if "/data/" in n.lower() or n.lower().startswith("data/")]
    pool = preferred or names
    if not pool:
        raise ValueError("No CSV found in the postcode directory zip")
    # The directory file is by far the largest CSV; pick it over any readme CSV.
    return max(pool, key=lambda n: zf.getinfo(n).file_size)


def _download_to_temp(url: str, suffix: str = "") -> str:
    """Stream a (large) download to a temp file and return its path.

    Used for the postcode directory, which is hundreds of MB: streaming to disk
    avoids holding the whole archive in memory.
    """
    import os
    import tempfile

    url = _require_url(url)
    fd, path = tempfile.mkstemp(suffix=suffix)
    with (
        httpx.Client(timeout=_TIMEOUT, follow_redirects=True) as client,
        client.stream("GET", url) as resp,
        os.fdopen(fd, "wb") as f,
    ):
        resp.raise_for_status()
        for chunk in resp.iter_bytes():
            f.write(chunk)
    return path


def load_postcodes(
    pool: ConnectionPool, url: str
) -> int:  # pragma: no cover - large file + DB COPY, exercised live
    """Load the ONS Postcode Directory (postcode to coordinate) for geocoding.

    Streams the directory CSV from the zip straight into a COPY, so the 2.6M row
    file never has to be held in memory at once.
    """
    import os

    path = _download_to_temp(url, suffix=".zip")
    n = 0
    lsoa_msoa: dict[str, str] = {}
    try:
        with (
            zipfile.ZipFile(path) as zf,
            pool.connection() as conn,
            conn.transaction(),
        ):
            member = _onspd_csv_member(zf)
            conn.execute("TRUNCATE postcode_centroid")
            with (
                zf.open(member) as raw,
                io.TextIOWrapper(raw, encoding="utf-8-sig", errors="ignore") as text,
                conn.cursor() as cur,
                cur.copy(
                    "COPY postcode_centroid (postcode, lat, lng) FROM STDIN"
                ) as copy,
            ):
                reader = csv.DictReader(text)
                lsoa_c, msoa_c = _lsoa_msoa_columns(reader.fieldnames or [])
                for row in reader:
                    parsed = _postcode_centroid(row)
                    if parsed is not None:
                        copy.write_row(parsed)
                        n += 1
                    # Build the LSOA to MSOA lookup as a side effect: the ONSPD
                    # carries both codes per postcode, so IMD can aggregate to
                    # MSOA without a separate lookup file.
                    if lsoa_c and msoa_c:
                        lsoa = str(row.get(lsoa_c) or "").strip()
                        msoa = str(row.get(msoa_c) or "").strip()
                        if lsoa and msoa:
                            lsoa_msoa.setdefault(lsoa, msoa)
            if lsoa_msoa:
                conn.execute("TRUNCATE lsoa_msoa")
                with conn.cursor() as cur:
                    cur.executemany(
                        "INSERT INTO lsoa_msoa (lsoa, msoa) VALUES (%s, %s) "
                        "ON CONFLICT (lsoa) DO NOTHING",
                        list(lsoa_msoa.items()),
                    )
    finally:
        os.unlink(path)
    return n


def _lsoa_msoa_columns(fieldnames: list[str]) -> tuple[str | None, str | None]:
    """Find the LSOA and MSOA (2011) code columns in an ONSPD/NSPL header."""
    return (
        t.find_column(fieldnames, ("lsoa11", "lsoa11cd", "lsoa")),
        t.find_column(fieldnames, ("msoa11", "msoa11cd", "msoa")),
    )


def _is_ods_url(url: str) -> bool:
    """True for an ODS ORD (Organisation Reference Data) API endpoint."""
    low = url.lower()
    return "spineservices" in low or "/ord/" in low


def _fetch_ods_organisations(url: str) -> list[dict]:
    """Page the ODS ORD organisations endpoint into one list.

    The sync API returns {"Organisations": [...]} and pages by Offset/Limit;
    each list item carries the ODS code (OrgId), Name and PostCode, which is all
    we need (coordinates come from geocoding the postcode).
    """
    parts = urlparse(_require_url(url))
    base = dict(parse_qsl(parts.query))
    limit = int(base.get("Limit", "1000"))
    results: list[dict] = []
    # The ODS API offset is 1-based (record number to start at); 0 is rejected.
    offset = 1
    # The ODS API gateway does content negotiation and rejects a non-browser
    # client: send a JSON Accept and a browser User-Agent, or it returns 406.
    headers = {
        "Accept": "application/json",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
    }
    with httpx.Client(
        timeout=_TIMEOUT, follow_redirects=True, headers=headers
    ) as client:
        for _ in range(10_000):  # safety bound
            params = {**base, "Limit": str(limit), "Offset": str(offset)}
            u = urlunparse(parts._replace(query=urlencode(params)))
            resp = client.get(u)
            if resp.status_code >= 400:
                # Surface the server's own message so a remaining failure is
                # diagnosable from the dataset status.
                raise ValueError(
                    f"ODS API returned {resp.status_code}. {resp.text[:300]}"
                )
            orgs = resp.json().get("Organisations") or []
            if not orgs:
                break
            results.extend(orgs)
            offset += len(orgs)
            if len(orgs) < limit:
                break
    return results


def _ods_hospital_points(pool: ConnectionPool, orgs: list[dict]) -> list[dict]:
    """ODS organisations to hospital points, geocoding postcodes via the
    postcode_centroid table loaded from the ONS Postcode Directory."""
    wanted: dict[str, list[tuple[str, str]]] = {}
    for o in orgs:
        pc = _norm_postcode(str(o.get("PostCode") or o.get("Postcode") or ""))
        if not pc:
            continue
        name = str(o.get("Name") or "").strip()
        code = str(o.get("OrgId") or o.get("OrgLink") or "").strip()
        wanted.setdefault(pc, []).append((name, code))
    if not wanted:
        raise ValueError("No postcodes in the ODS response to geocode")
    coords: dict[str, tuple[float, float]] = {}
    with pool.connection() as conn:
        rows = conn.execute(
            "SELECT postcode, lat, lng FROM postcode_centroid WHERE postcode = ANY(%s)",
            [list(wanted)],
        ).fetchall()
        for pc, lat, lng in rows:
            coords[pc] = (lat, lng)
    if not coords:
        raise ValueError(
            "No hospital postcodes matched. Load the Postcodes dataset first so "
            "ODS postcodes can be geocoded."
        )
    points: list[dict] = []
    for pc, entries in wanted.items():
        if pc not in coords:
            continue
        lat, lng = coords[pc]
        for name, code in entries:
            points.append(
                {"name": name or None, "org_code": code or None, "lat": lat, "lng": lng}
            )
    return points


def _hospital_points(records: list[dict]) -> list[dict]:
    """Hospitals as {name, org_code, lng, lat} from an NHS sites file.

    Accepts lat/long columns directly, or easting/northing (BNG) which we
    reproject. Org code (parent/trust ODS code) is kept for the waiting-times
    join when present.
    """
    if not records:
        return []
    f = list(records[0].keys())
    lat_c = t.find_column(f, ("latitude", "lat"))
    lng_c = t.find_column(f, ("longitude", "long", "lng"))
    east_c = t.find_column(f, ("easting",))
    north_c = t.find_column(f, ("northing",))
    name_c = t.find_column(
        f, ("organisationname", "hospital name", "name", "site name")
    )
    org_c = t.find_column(
        f,
        (
            "parentodscode",
            "parent organisation code",
            "trust code",
            "organisationcode",
            "org code",
        ),
    )
    points: list[dict] = []
    for r in records:
        lat = lng = None
        if lat_c and lng_c:
            lat = t.parse_number(r.get(lat_c))
            lng = t.parse_number(r.get(lng_c))
        elif east_c and north_c:
            e = t.parse_number(r.get(east_c))
            n = t.parse_number(r.get(north_c))
            if e and n:
                lng, lat = _bng_to_wgs(e, n)
        if lat is None or lng is None or (lat == 0 and lng == 0):
            continue
        points.append(
            {
                "name": str(r.get(name_c) or "").strip() if name_c else None,
                "org_code": str(r.get(org_c) or "").strip() if org_c else None,
                "lat": lat,
                "lng": lng,
            }
        )
    return points


def load_hospitals(
    pool: ConnectionPool, url: str, area_type: str = "MSOA"
) -> int:  # pragma: no cover - PostGIS spatial join, exercised live
    # ODS ORD API carries org code + postcode (no coordinates), so geocode via
    # the postcode directory; a plain NHS sites CSV carries coordinates directly.
    if _is_ods_url(url):
        points = _ods_hospital_points(pool, _fetch_ods_organisations(url))
    else:
        points = _hospital_points(_load_records(url))
    if not points:
        return 0
    with pool.connection() as conn, conn.transaction():
        conn.execute("TRUNCATE hospital")
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT INTO hospital (name, org_code, lat, lng, geom) "
                "VALUES (%s, %s, %s, %s, ST_SetSRID(ST_MakePoint(%s, %s), 4326))",
                [
                    (p["name"], p["org_code"], p["lat"], p["lng"], p["lng"], p["lat"])
                    for p in points
                ],
            )
        # Per-MSOA distance to the nearest hospital, in km (great-circle).
        conn.execute("DELETE FROM area_metric WHERE metric_key = 'hospital_km'")
        conn.execute(
            "INSERT INTO area_metric (area_code, area_type, metric_key, value) "
            "SELECT b.area_code, %s, 'hospital_km', round((("
            "  SELECT ST_Distance(ST_Centroid(b.geom)::geography, h.geom::geography) "
            "  FROM hospital h ORDER BY ST_Centroid(b.geom) <-> h.geom LIMIT 1"
            ") / 1000.0)::numeric, 1) "
            "FROM geo_boundaries b WHERE b.area_type = %s",
            [area_type, area_type],
        )
    return len(points)


def _crime_points_stream(
    text: object,
) -> Iterator[tuple[float, float]]:  # pragma: no cover - exercised live
    """Yield (lng, lat) from a crime CSV stream, skipping members without coords."""
    reader = csv.DictReader(text)  # type: ignore[arg-type]
    fields = reader.fieldnames or []
    lng_c = t.find_column(fields, ("longitude", "long", "lng"))
    lat_c = t.find_column(fields, ("latitude", "lat"))
    if lng_c is None or lat_c is None:
        return
    for r in reader:
        lng = t.parse_number(r.get(lng_c))
        lat = t.parse_number(r.get(lat_c))
        if lng is not None and lat is not None:
            yield (lng, lat)


def _iter_crime_points_from_path(
    path: str, name: str
) -> Iterator[tuple[float, float]]:  # pragma: no cover - live
    """Stream (lng, lat) from a crime file on disk: every CSV in a zip (the
    data.police.uk archive), or a single CSV. Streaming keeps a multi-GB national
    archive off the heap."""
    if name.lower().endswith(".zip"):
        with zipfile.ZipFile(path) as zf:
            for member in zf.namelist():
                if not member.lower().endswith(".csv"):
                    continue
                with (
                    zf.open(member) as raw,
                    io.TextIOWrapper(raw, encoding="utf-8-sig", errors="ignore") as text,
                ):
                    yield from _crime_points_stream(text)
    else:
        with open(path, encoding="utf-8-sig", errors="ignore") as text:
            yield from _crime_points_stream(text)


def _load_crime_from_path(
    pool: ConnectionPool, path: str, name: str, area_type: str
) -> int:  # pragma: no cover - PostGIS spatial join, exercised live
    """Stream crime points from a file into a staging table, then aggregate to
    crime per 1,000 residents per area in PostGIS. Existing crime is replaced
    only when the file yields points, so a bad file does not wipe the metric."""
    with pool.connection() as conn, conn.transaction():
        conn.execute("CREATE TEMP TABLE _crime (lng float8, lat float8) ON COMMIT DROP")
        n_points = 0
        with (
            conn.cursor() as cur,
            cur.copy("COPY _crime (lng, lat) FROM STDIN") as copy,
        ):
            for lng, lat in _iter_crime_points_from_path(path, name):
                copy.write_row((lng, lat))
                n_points += 1
        if n_points == 0:
            return 0
        conn.execute("DELETE FROM area_metric WHERE metric_key = 'crime_per_1k'")
        conn.execute(
            "INSERT INTO area_metric (area_code, area_type, metric_key, value) "
            "SELECT b.area_code, %s, 'crime_per_1k', "
            "round(1000.0 * count(*) / NULLIF(d.population, 0)) "
            "FROM _crime c JOIN geo_boundaries b "
            "ON b.area_type = %s AND ST_Contains("
            "b.geom, ST_SetSRID(ST_MakePoint(c.lng, c.lat), 4326)) "
            "JOIN census_demographics d ON d.area_code = b.area_code "
            "WHERE d.population > 0 "
            "GROUP BY b.area_code, d.population",
            [area_type, area_type],
        )
        n = conn.execute(
            "SELECT count(*) FROM area_metric WHERE metric_key = 'crime_per_1k'"
        ).fetchone()[0]
    return int(n)


def load_crime(
    pool: ConnectionPool, url: str, area_type: str = "MSOA"
) -> int:  # pragma: no cover - PostGIS spatial join, exercised live
    import os

    suffix = ".zip" if url.lower().endswith(".zip") else ".csv"
    path = _download_to_temp(url, suffix=suffix)
    try:
        return _load_crime_from_path(pool, path, url, area_type)
    finally:
        os.unlink(path)


def load_crime_file(
    pool: ConnectionPool, path: str, name: str, area_type: str = "MSOA"
) -> int:  # pragma: no cover - PostGIS spatial join, exercised live
    """Load crime from an uploaded file on disk (CSV or data.police.uk zip)."""
    return _load_crime_from_path(pool, path, name, area_type)


def _replace_table(
    pool: ConnectionPool,
    table: str,
    columns: tuple[str, ...],
    rows: list[dict],
    source: str,
    geometry_columns: dict[str, str] | None = None,
) -> int:
    geometry_columns = geometry_columns or {}
    col_list = ", ".join(columns)
    placeholders = ", ".join(geometry_columns.get(c, "%s") for c in columns)
    insert_sql = f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"
    params = [[row.get(c) for c in columns] for row in rows]
    with pool.connection() as conn, conn.transaction():
        conn.execute(f"TRUNCATE {table}")
        with conn.cursor() as cur:
            # Batched insert: far faster than per-row for thousands of areas.
            cur.executemany(insert_sql, params)
        conn.execute(
            "INSERT INTO reference_load (table_name, source, source_version) "
            "VALUES (%s, %s, %s)",
            [table, source, "downloaded"],
        )
    return len(rows)


# --- Public load functions ---------------------------------------------------


def load_boundaries(pool: ConnectionPool, url: str, area_type: str = "MSOA") -> int:
    keys = _BOUNDARY_KEYS[area_type]
    geojson = fetch_arcgis_geojson(url, out_fields=f"{keys['code']},{keys['name']}")
    rows = _boundary_rows(geojson, area_type)
    return _replace_table(
        pool,
        "geo_boundaries",
        ("area_code", "area_type", "area_name", "geom"),
        rows,
        source="ONS Open Geography Portal",
        geometry_columns={"geom": "ST_Multi(ST_SetSRID(ST_GeomFromGeoJSON(%s), 4326))"},
    )


def load_demographics(
    pool: ConnectionPool, age_url: str, households_url: str, area_type: str = "MSOA"
) -> int:
    rows = _demographics_rows(
        _fetch_csv_text(age_url, area_type),
        _fetch_csv_text(households_url, area_type),
        area_type,
    )
    # A household file that parses to no totals nulls households, family share
    # and the addressable segments silently. Fail loudly instead so the cause is
    # visible on the Reference data page rather than surfacing as empty cards.
    if households_url and rows and not any(r["households"] is not None for r in rows):
        raise ValueError(
            "Parsed zero household counts from the household-composition file. "
            "Check the household URL points to the ONS TS003 bulk CSV: its total "
            "column should read 'Household composition: Total'."
        )
    return _replace_table(
        pool,
        "census_demographics",
        (
            "area_code",
            "area_type",
            "population",
            "households",
            "age_0_15",
            "age_16_34",
            "age_35_54",
            "age_55_74",
            "age_75_plus",
            "median_age",
            "family_household_share",
        ),
        rows,
        source="ONS Census 2021",
    )


def load_tenure(pool: ConnectionPool, url: str, area_type: str = "MSOA") -> int:
    rows = _tenure_rows(_fetch_csv_text(url, area_type), area_type)
    if rows and not any(
        any(
            r[k] is not None
            for k in (
                "owns_outright",
                "owns_with_mortgage",
                "social_rented",
                "private_rented",
            )
        )
        for r in rows
    ):
        raise ValueError(
            "Parsed zero tenure shares from the tenure file. Check the URL points "
            "to the ONS TS054 bulk CSV: its total column should read 'Tenure of "
            "household: Total'."
        )
    return _replace_table(
        pool,
        "census_tenure",
        (
            "area_code",
            "area_type",
            "owns_outright",
            "owns_with_mortgage",
            "social_rented",
            "private_rented",
        ),
        rows,
        source="ONS Census 2021",
    )


def _wkt_point(value: object) -> tuple[float | None, float | None]:
    """Longitude, latitude from a WKT point like 'POINT (-1.23 52.10)'."""
    m = re.search(r"POINT\s*\(\s*(-?\d+\.?\d*)\s+(-?\d+\.?\d*)\s*\)", str(value or ""))
    if not m:
        return None, None
    return float(m.group(1)), float(m.group(2))


def _to_int(value: float | None) -> int | None:
    return None if value is None else int(round(value))


def _development_site_rows(text: str) -> list[dict]:
    """Brownfield land register rows as development site points.

    Reads the planning.data.gov.uk brownfield-land CSV: a 'point' column carries
    the WKT centroid, with hectares and minimum/maximum net dwellings. Tolerant
    of column-name variants so a plain lat/long file also parses.
    """
    records = _csv_rows(text)
    if not records:
        return []
    f = list(records[0].keys())
    point_c = t.find_column(f, ("point",))
    lat_c = t.find_column(f, ("latitude", "lat"))
    lng_c = t.find_column(f, ("longitude", "long", "lng"))
    ref_c = t.find_column(f, ("reference", "site reference", "slug"))
    name_c = t.find_column(f, ("name", "site-address", "site address", "address"))
    ha_c = t.find_column(f, ("hectares", "site area", "area"))
    maxd_c = t.find_column(
        f, ("maximum-net-dwellings", "maximum net dwellings", "max dwellings")
    )
    mind_c = t.find_column(
        f, ("minimum-net-dwellings", "minimum net dwellings", "min dwellings")
    )
    rows: list[dict] = []
    for r in records:
        lng = lat = None
        if point_c:
            lng, lat = _wkt_point(r.get(point_c))
        if (lat is None or lng is None) and lat_c and lng_c:
            lat = t.parse_number(r.get(lat_c))
            lng = t.parse_number(r.get(lng_c))
        if lat is None or lng is None or (lat == 0 and lng == 0):
            continue
        rows.append(
            {
                "reference": (str(r.get(ref_c)).strip() if ref_c else None) or None,
                "name": (str(r.get(name_c)).strip() if name_c else None) or None,
                "hectares": t.parse_number(r.get(ha_c)) if ha_c else None,
                "min_dwellings": _to_int(t.parse_number(r.get(mind_c)))
                if mind_c
                else None,
                "max_dwellings": _to_int(t.parse_number(r.get(maxd_c)))
                if maxd_c
                else None,
                "lat": lat,
                "lng": lng,
            }
        )
    return rows


def load_development_sites(
    pool: ConnectionPool, url: str, area_type: str = "MSOA"
) -> int:  # pragma: no cover - PostGIS insert, exercised live
    rows = _development_site_rows(_fetch_csv_text(url, area_type))
    if not rows:
        raise ValueError(
            "Parsed zero development sites. Check the URL points to the "
            "planning.data.gov.uk brownfield-land CSV (it should carry a 'point' "
            "column, or latitude and longitude)."
        )
    with pool.connection() as conn, conn.transaction():
        conn.execute("TRUNCATE development_site")
        with conn.cursor() as cur:
            cur.executemany(
                "INSERT INTO development_site "
                "(reference, name, hectares, min_dwellings, max_dwellings, lat, "
                "lng, geom) VALUES "
                "(%s, %s, %s, %s, %s, %s, %s, ST_SetSRID(ST_MakePoint(%s, %s), 4326))",
                [
                    (
                        r["reference"],
                        r["name"],
                        r["hectares"],
                        r["min_dwellings"],
                        r["max_dwellings"],
                        r["lat"],
                        r["lng"],
                        r["lng"],
                        r["lat"],
                    )
                    for r in rows
                ],
            )
    return len(rows)


def load_income(pool: ConnectionPool, url: str, area_type: str = "MSOA") -> int:
    if url.lower().endswith((".xlsx", ".xlsm")):
        records = _read_xlsx(_get_bytes(url))
    else:
        records, _ = _read_csv(_get_text(url))
    rows = _income_rows(records, area_type)
    return _replace_table(
        pool,
        "income_estimates",
        ("area_code", "area_type", "median_income", "mean_income"),
        rows,
        source="ONS income estimates",
    )


def load_house_prices(pool: ConnectionPool, url: str, area_type: str = "MSOA") -> int:
    # Fetch once; if the user pasted the ONS dataset page, follow its download
    # link to the real file. Then parse by the resolved file's extension.
    data = _get_bytes(url)
    resolved = _resolve_data_url(url, data)
    if resolved:
        url = resolved
        data = _get_bytes(url)
    low = url.lower()
    if low.endswith(".xls"):
        records = _read_xls(data)
    elif low.endswith((".xlsx", ".xlsm")):
        records = _read_xlsx(data)
    else:
        records, _ = _read_csv(data.decode("utf-8-sig", "ignore"))
    rows = _house_price_rows(records, area_type)
    return _replace_table(
        pool,
        "house_prices",
        ("area_code", "area_type", "median_price"),
        rows,
        source="ONS HPSSA",
    )
