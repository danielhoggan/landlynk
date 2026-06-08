"""Income loader: ONS model-based income estimates for small areas.

Source is the ONS "Income estimates for small areas, England and Wales" release
(Open Government Licence), MSOA level, published as a spreadsheet. Mean and
median net annual household income are read by column, configurable because the
sheet layout changes between releases. Suppressed cells stay None.

Reads .xlsx via openpyxl or .csv directly. The transform over parsed rows is
pure and unit tested.
"""

from __future__ import annotations

import csv

from .base import ReferenceLoader
from .transforms import parse_number

# Substrings used to find the columns when explicit names are not given.
_AREA_CODE_NEEDLES = ("msoa code", "area code", "geography code", "code")
_MEAN_NEEDLES = ("net annual income (mean)", "mean")
_MEDIAN_NEEDLES = ("net annual income (median)", "median")


def _find_column(fieldnames: list[str], needles: tuple[str, ...]) -> str | None:
    lowered = [(f, f.lower()) for f in fieldnames]
    for needle in needles:
        for original, low in lowered:
            if needle in low:
                return original
    return None


class IncomeLoader(ReferenceLoader):
    target_table = "income_estimates"
    columns = ("area_code", "area_type", "median_income", "mean_income")

    def __init__(
        self,
        spec,
        source: str,
        area_type: str = "MSOA",
        area_code_col: str | None = None,
        mean_col: str | None = None,
        median_col: str | None = None,
        db=None,
    ) -> None:
        super().__init__(spec, source, db)
        self.area_type = area_type
        self.area_code_col = area_code_col
        self.mean_col = mean_col
        self.median_col = median_col

    def fetch(self) -> list[dict]:
        if self.source.lower().endswith((".xlsx", ".xlsm")):
            return self._read_xlsx(self.source)
        with open(self.source, newline="", encoding="utf-8-sig") as fh:
            return list(csv.DictReader(fh))

    @staticmethod
    def _read_xlsx(path: str) -> list[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
        records: list[dict] = []
        for values in rows_iter:
            records.append(dict(zip(header, values, strict=False)))
        return records

    def transform(self, raw: list[dict]) -> list[dict]:
        if not raw:
            return []
        fieldnames = list(raw[0].keys())
        code_col = self.area_code_col or _find_column(fieldnames, _AREA_CODE_NEEDLES)
        mean_col = self.mean_col or _find_column(fieldnames, _MEAN_NEEDLES)
        median_col = self.median_col or _find_column(fieldnames, _MEDIAN_NEEDLES)
        if code_col is None:
            raise ValueError(f"No area code column found in {fieldnames}")

        rows: list[dict] = []
        for record in raw:
            code = str(record.get(code_col) or "").strip()
            if not code:
                continue
            rows.append(
                {
                    "area_code": code,
                    "area_type": self.area_type,
                    "median_income": (
                        parse_number(record.get(median_col)) if median_col else None
                    ),
                    "mean_income": (
                        parse_number(record.get(mean_col)) if mean_col else None
                    ),
                }
            )
        return rows
